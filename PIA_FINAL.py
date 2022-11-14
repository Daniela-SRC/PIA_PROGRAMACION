import csv
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import os
from os import path
import sqlite3
from sqlite3 import Error
import sys
fecha_actual = datetime.date.today()
dia_actual = fecha_actual.day
mes_actual = fecha_actual.month
año_actual = fecha_actual.year
tupla_actual = (dia_actual, mes_actual, año_actual)
opcion = 0
clave_cliente = 0
clave_sala = 0
clave_registro = 0
fechaExistente = False
FechaValida = True
row=2

def menu():
    Correcto = True
    try:
        opc = int(input("Menú Principal\n" +
                        "Seleccione la opcion que guste:\n"+
                        "1.- Reservar\n" +
                        "2.- Reportes\n" +
                        "3.- Registrar un nuevo cliente\n" +
                        "4.- Registrar una sala\n" +
                        "5.- Finalizar\n"))
    except ValueError:
        print("INGRESE UN NUMERO!")
        Correcto = False
    if Correcto:
        if opc < 1 or opc > 5:
            print("INGRESE UN NUMERO DEL 1 AL 5")
        return opc

def menu_reserva():
    Correcto = True
    try:
        opc1 = int(input("\nSeleccione la opcion que guste:\n"+
                        "1.- Registrar una reservación\n" +
                        "2.- Editar el nombre de un evento reservado\n" +
                        "3.- Consultar disponibilidad de salas\n" +
                        "4.- Eliminar una reservacion\n" +
                        "5.- Volver al menu principal\n"))
    except ValueError:
        print("INGRESE UN NUMERO!")
        Correcto = False
    if Correcto:
        if opc1 < 1 or opc1 > 5:
            print("INGRESE UN NUMERO DEL 1 AL 5")
        return opc1  

def menu_reporte():
     Correcto = True
     try:
        opc2 = int(input("\nSeleccione la opcion que guste:\n"+
                        "1.- Reporte de reservaciones para una fecha\n" +
                        "2.- Exportar reporte a Excel\n" +
                        "3.- Volver al menu principal\n"))
     except ValueError:
        print("INGRESE UN NUMERO!")
        Correcto = False
     if Correcto:
        if opc2 < 1 or opc2 > 3:
            print("INGRESE UN NUMERO DEL 1 AL 3")
        return opc2  


abrir_bd = False
try:
    existe = open("bd_eventos.db")
    print("La base de datos ya existe")
    existe.close()
except FileNotFoundError:
    print("No se encontro la base de datos, a continuacion sera creada")
    try:
        with sqlite3.connect("bd_eventos.db", detect_types = sqlite3.PARSE_DECLTYPES|sqlite3.PARSE_COLNAMES) as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("CREATE TABLE IF NOT EXISTS clientes (clave_c INTEGER PRIMARY KEY, nombre TEXT NOT NULL, apellidos TEXT NOT NULL);")
            mi_cursor.execute("CREATE TABLE IF NOT EXISTS salas (clave_s INTEGER PRIMARY KEY, nombre TEXT NOT NULL, cupo INTEGER NOT NULL);")
            mi_cursor.execute("CREATE TABLE IF NOT EXISTS turnos (clave_turno INTEGER PRIMARY KEY, nombre_turno TEXT NOT NULL);")
            mi_cursor.execute("CREATE TABLE IF NOT EXISTS reservaciones (clave INTEGER PRIMARY KEY,nombre TEXT NOT NULL, turno TEXT NOT NULL,fecha_ev timestamp NOT NULL,cve_sala INTEGER NOT NULL,cve_cliente INTEGER NOT NULL,FOREIGN KEY(cve_sala) REFERENCES salas(clave_s),FOREIGN KEY(cve_cliente) REFERENCES clientes(clave_c));")
            
            mi_cursor.execute("INSERT INTO turnos VALUES(1, 'Matutino')")
            mi_cursor.execute("INSERT INTO turnos VALUES(2, 'Vespertino')")
            mi_cursor.execute("INSERT INTO turnos VALUES(3, 'Nocturno')")
    except Error as e:
        print (e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    abrir_bd = True
finally:      
    while opcion !=5:
         if opcion == 1:
            opcion1 = menu_reserva()
            if opcion1 == 1:
                registroExistente = True
                try:
                    with sqlite3.connect("bd_eventos.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT * from clientes")
                        mi_cursor.execute("SELECT * from salas")
                        registro1 = mi_cursor.fetchall()
                        registro = mi_cursor.fetchall()
                        if not registro1 or registro:
                            print("No se pueden realizar reservaciones sin clientes o salas registradas")
                            registroExistente = False
                except Error as e:
                        print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                    conn.close()
                if registroExistente:
                    print("Registrar la reservación de una sala para un evento\n")
                    try:
                        with sqlite3.connect("bd_eventos.db") as conn:
                            mi_cursor = conn.cursor()
                            mi_cursor.execute("SELECT * from clientes")
                            clientes = mi_cursor.fetchall()
                            print("Clientes existentes")
                            print("{:<20} {:<20} {:<20}".format('Clave Cliente','Nombre','Apellido'))
                            for clave_c, nombre, apellido in clientes:
                                print("{:<20} {:<20} {:<20}".format(clave_c, nombre, apellido))
                    except Error as e:
                            print (e)
                    except:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                    finally:
                        conn.close()
                                        
                    while True:
                        cve_correcto = True
                        try:                        
                            cve_cliente=int(input("Ingrese el ID del cliente: "))
                            try:
                                with sqlite3.connect("bd_eventos.db") as conn:
                                    mi_cursor = conn.cursor()
                                    valores = {"clave_cliente": cve_cliente}
                                    mi_cursor.execute("SELECT clave_c from clientes where clave_c = :clave_cliente", valores)
                                    registro = mi_cursor.fetchall()
                                    if registro:                            
                                        try:
                                            with sqlite3.connect("bd_eventos.db") as conn:
                                                mi_cursor = conn.cursor()
                                                mi_cursor.execute("SELECT * from salas")
                                                salas = mi_cursor.fetchall()
                                                print("Salas existentes")
                                                print("{:<20} {:<20} {:<20}".format('Clave Sala','Sala','Cupo'))
                                                for clave_s, sala, cupo_s in salas:
                                                    print("{:<20} {:<20} {:<20}".format(clave_s, sala, cupo_s))

                                        except Error as e:
                                                print (e)
                                        except:
                                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                        finally:
                                            conn.close()
                                        while True:
                                            cve_correcto = True
                                            try:                        
                                                cve_sala=int(input("Ingrese el ID de la sala: "))
                                                try:
                                                    with sqlite3.connect("bd_eventos.db") as conn:
                                                        mi_cursor = conn.cursor()
                                                        valores = {"clave_sala": cve_sala}
                                                        mi_cursor.execute("SELECT clave_s from salas where clave_s = :clave_sala", valores)
                                                        registro = mi_cursor.fetchall()
                                                        if registro:                                        
                                                            while True:
                                                                FechaValida = True
                                                                while FechaValida:
                                                                    try:
                                                                        fecha_reservada = input("Ingrese la fecha que desea reservar (dd/mm/aaaa): ")
                                                                        fecha_procesada = datetime.datetime.strptime(fecha_reservada,"%d/%m/%Y").date()
                                                                        fecha_guardar = datetime.date(fecha_procesada.year, fecha_procesada.month, fecha_procesada.day)
                                                                        dia_reservado = fecha_procesada.day
                                                                        mes_reservado = fecha_procesada.month
                                                                        año_reservado = fecha_procesada.year
                                                                    except ValueError:
                                                                        print("LA FECHA INTRODUCIDA NO ES VALIDA")
                                                                        
                                                                    else:
                                                                        FechaValida = False
                                                                
                                                                if True:
                                                                    
                                                                    fechaActual = datetime.date.today()                                                
                                                                    fechaValida = fechaActual + datetime.timedelta(days = 2)
                                                                    if fechaValida > fecha_procesada:
                                                                        print("Para reservar una fecha debe hacerlo con 2 dias de anticipación")
                                                                    else:                                                    
                                                                        while True:
                                                                            horario_evento = input("Ingrese el numero (1,2 ó 3) del horario del evento que desee (1.-MATUTINO, 2.-VESPERTINO, 3.-NOCTURNO): ")                                                
                                                                            if horario_evento == "1":
                                                                                horario_evento="Matutino"
                                                                                break
                                                                                
                                                                            elif horario_evento == "2":
                                                                                horario_evento="Vespertino"
                                                                                break
                                                                                
                                                                            elif horario_evento == "3":
                                                                                horario_evento="Nocturno"
                                                                                break
                                                                            else:
                                                                                print("Tipo de dato incorrecto favor de ingresar un numero del 1 al 3")                                                    
                                                                        try:
                                                                            with sqlite3.connect("bd_eventos.db") as conn:
                                                                                mi_cursor = conn.cursor()
                                                                                valores = {"clave_sala": cve_sala,"turno": horario_evento, "fecha_guardar": fecha_guardar}                                                                
                                                                                mi_cursor.execute("SELECT *FROM reservaciones WHERE cve_sala = :clave_sala AND turno = :turno AND fecha_ev = :fecha_guardar", valores)                                                                
                                                                                registro = mi_cursor.fetchall()                                                                
                                                
                                                                                if registro:
                                                                                    print("Lo sentimos, empalma sala o turno ingresado, favor de verificar")
                                                                                    break
                                                                                else:
                                                                                    while True:
                                                                                        nombre_evento=input("Ingresa el nombre del evento: ")
                                                                                        nombre_evento_sinespacio = nombre_evento.strip()
                                                                                        if nombre_evento_sinespacio == "":
                                                                                            print("El nombre del evento no se puede omitir")
                                                                                        else:
                                                                                            break
                                                                                    try:
                                                                                        with sqlite3.connect("bd_eventos.db", detect_types = sqlite3.PARSE_DECLTYPES|sqlite3.PARSE_COLNAMES) as conn:
                                                                                            mi_cursor = conn.cursor()
                                                                                            valores = {"nombre": nombre_evento, "turno": horario_evento, "fecha_ev": fecha_procesada, "clave_s": cve_sala, "cve_cliente": cve_cliente}
                                                                                            mi_cursor.execute("INSERT INTO reservaciones (nombre, turno, fecha_ev, cve_sala, cve_cliente) VALUES(:nombre, :turno, :fecha_ev, :clave_s, :cve_cliente)", valores)
                                                                                            print("La reservación ha sido éxitosa\n")
                                                                                            
                                                                                    except Error as e:
                                                                                        print (e)
                                                                                    except:
                                                                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                                                                    finally:
                                                                                        conn.close()
                                                                        except Error as e:
                                                                            print (e)
                                                                        except:
                                                                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                                                        finally:
                                                                            conn.close()
                                                                        break                                                    
                                                                                                                
                                                        else:
                                                            print(f"No existe la sala ingresada, intente de nuevo\n")
                                                            cve_correcto =False
                                                except Error as e:
                                                    print (e)
                                                except:
                                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                                finally:
                                                    conn.close()                            
                                    
                                            except ValueError:
                                                print("Ingrese un numero")
                                                cve_correcto = False
                                            if cve_correcto:
                                                break
                                    else:
                                            print(f"No existe el cliente registrado, intente de nuevo\n")
                                            cve_correcto = False
                            except Error as e:
                                print (e)
                            except:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                            finally:
                                conn.close()
                        except ValueError:
                            print("Ingrese un numero")
                            cve_correcto = False
                        if cve_correcto:
                            break
                                
            if opcion1 == 2:
                print("Editar el nombre de un evento reservado\n")
                try:
                    with sqlite3.connect("bd_eventos.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT clave, nombre from reservaciones")
                        reservacion = mi_cursor.fetchall()
                        print("Reservacion")
                        print("{:<20} {:<20}".format('Clave Reservacion','Evento'))
                        for clave_r, evento in reservacion:
                            print("{:<20} {:<20}".format(clave_r, evento))
                except Error as e:
                        print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                    conn.close()
                    
                while True:
                    cve_correcto = True
                    try:                        
                        editar=int(input("\nIngrese el ID de la reservación que gusta modificar: "))
                        try:
                            with sqlite3.connect("bd_eventos.db") as conn:
                                detect_types = sqlite3.PARSE_DECLTYPES|sqlite3.PARSE_COLNAMES
                                mi_cursor = conn.cursor()
                                valores={"clave":editar}
                                mi_cursor.execute("SELECT * FROM reservaciones WHERE clave = :clave", valores)
                                registros = mi_cursor.fetchall()
                                if registros:
                                    while True:
                                        nombre_ev=input("Ingresa el nuevo nombre del evento: ")
                                        nombre_evento_sinespacio = nombre_ev.strip()
                                        if nombre_evento_sinespacio == "":
                                            print("El nombre del evento no se puede omitir")
                                        else:
                                            try:
                                                with sqlite3.connect("bd_eventos.db") as conn:
                                                    mi_cursor = conn.cursor()
                                                    valores={"clave":editar, "nombre":nombre_ev}
                                                    mi_cursor.execute("UPDATE reservaciones SET nombre= :nombre WHERE clave = :clave", valores)
                                                    conn.commit()
                                                    print("Registro editado exitosamente\n")
                                            except Error as e:
                                                print (e)
                                            except:
                                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                            finally:
                                                conn.close()
                                            break
                                else:
                                    print("La clave de la reservacion no existe, intente nuevamente\n")
                                    cve_correcto = False
                        except Error as e:
                            print (e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            conn.close()
                    except ValueError:
                        print("Ingrese un numero")
                        cve_correcto = False
                    if cve_correcto:
                        break
                
                    
            if opcion1 == 3:
                print('Disponibilidad de salas\n')
                ocupadas = set()
                posibles = set()
                while True:
                  try: 
                      fecha_consulta = input("Ingrese la fecha que desea consultar (dd/mm/aa): ")
                      fecha_consulta = datetime.datetime.strptime(fecha_consulta, "%d/%m/%Y").date()
                      if fecha_consulta == "":
                        print("")
                        print("No se puede omitir")
                        continue
                      else:
                        with sqlite3.connect("bd_eventos.db") as conn:
                          cursor = conn.cursor()
                          cursor.execute("SELECT fecha_ev, turno, cve_sala FROM reservaciones")
                          registros_reservaciones = cursor.fetchall()
                          cursor.execute("SELECT clave_turno, nombre_turno FROM turnos")
                          registros_turnos = cursor.fetchall()
                          cursor.execute("SELECT clave_s, nombre, cupo FROM salas")
                          registros_salas = cursor.fetchall()

                          if registros_reservaciones:
                            for fecha, turno, clave_sala in registros_reservaciones:
                                fecha_recuperada = fecha
                                fecha_recuperada_dt = datetime.datetime.strptime(fecha_recuperada,'%Y-%m-%d').date()
                                if fecha_recuperada_dt == fecha_consulta:
                                    ocupadas.add((clave_sala, turno))
                            for clave , nombre_sala, cupo_sala in registros_salas:
                                for clave_turno, turnos in registros_turnos:
                                    posibles.add((clave, turnos))
                            
                          disponibles = sorted(posibles - ocupadas)
                          print("")

                          print(f"Salas disponibles para la fecha: {fecha_consulta}\n")
                          print("{:<20} {:<20}".format('Clave Sala','Turno'))
                          if registros_salas:
                              for clave_s, turno_s in disponibles:
                                print("{:<20} {:<20}".format(clave_s, turno_s))
                              break
                  except ValueError:
                    print("")
                    print("Formato de la fecha es incorrecta")
                    continue     
                  except Error as e:
                    print (e)
                  except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                    break
                  finally:
                    conn.close()
                        

                
            if opcion1 == 4:
                print("Eliminar una reservacion\n")
                try:
                    with sqlite3.connect("bd_eventos.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT * from reservaciones")
                        reservacion = mi_cursor.fetchall()
                        print("Reservaciones")
                        print("{:<20} {:<20} {:<20} {:<20} {:<20} {:<20}".format('Clave Reservacion','Evento', 'Turno', 'Fecha', 'Clave Sala', 'Clave Cliente'))
                        for clave_r, evento, turno_r, fecha_r, clave_sr, clave_cr in reservacion:
                            print("{:<20} {:<20} {:<20} {:<20} {:<20} {:<20}".format(clave_r, evento, turno_r, fecha_r, clave_sr, clave_cr))
                except Error as e:
                        print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                    conn.close()
                    
                while True:
                    cve_correcto = True
                    try:                        
                        eliminar=int(input("\nIngrese el ID de la reservacion que quiera eliminar: "))
                        try:
                            with sqlite3.connect("bd_eventos.db") as conn:
                                detect_types = sqlite3.PARSE_DECLTYPES|sqlite3.PARSE_COLNAMES
                                mi_cursor = conn.cursor()
                                valores={"clave":eliminar}
                                mi_cursor.execute("SELECT * FROM reservaciones WHERE clave = :clave", valores)
                                registros = mi_cursor.fetchall()
                                if registros:
                                    print("\nRegistro a eliminar")
                                    print("{:<20} {:<20} {:<20} {:<20} {:<20} {:<20}".format('Clave Reservacion','Evento', 'Turno', 'Fecha', 'Clave Sala', 'Clave Cliente'))
                                    print("-" * 100)
                                    for clave_r, evento, turno_r, fecha_r, clave_sr, clave_cr in registros:
                                        print("{:<20} {:<20} {:<20} {:<20} {:<20} {:<20}".format(clave_r, evento, turno_r, fecha_r, clave_sr, clave_cr))
                                        fecha_date = fecha_r
                                        fecha_dt = datetime.datetime.strptime(fecha_date,'%Y-%m-%d').date()
                                        fecha_hoy = datetime.date.today()
                                        nueva_fecha = fecha_hoy + datetime.timedelta(days=2)
                                        confirmacion= int(input("¿Está seguro de eliminar estos datos? (1.- Si/ 2.- No)\n"))
                                        if confirmacion == 1:
                                            if nueva_fecha<fecha_dt:
                                                mi_cursor = conn.cursor()
                                                valores={"clave":eliminar}
                                                mi_cursor.execute("DELETE FROM reservaciones WHERE  clave = :clave", valores)
                                                registros = mi_cursor.fetchall()
                                                conn.commit()
                                                print("Registro eliminado\n")
                                                break
                                            else:
                                                print("Para eliminar un registro debe hacerlo 3 dias antes del evento\n")
                                                break
                                else:
                                    print("La clave de la reservacion no existe, intente de nuevo")
                                    cve_correcto = False
                        except Error as e:
                            print (e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            conn.close()
                    except ValueError:
                        print("Ingrese un numero")
                        cve_correcto = False
                    if cve_correcto:
                        break
                
                
                
            if opcion1 == 5:
                opcion = 0;
                print("-"*40)
                
         else:            
            if opcion == 2:
                opcion2 = menu_reporte()
                if opcion2 == 1:
                    print("Consulta de reservaciones\n")
                    #fechaExistente=False
                    FechaValida = True
                    while FechaValida:
                        try:
                            fecha_consulta = input("Ingrese la fecha que desea consultar (dd/mm/aaaa): ")
                            fecha_valida = datetime.datetime.strptime(fecha_consulta,"%d/%m/%Y").date()
                            dia_consulta = fecha_valida.day
                            mes_consulta = fecha_valida.month
                            año_consulta = fecha_valida.year
                        except ValueError:
                            print("Ingrese la fecha en formato dd/mm/aaaa")
                        else:
                            FechaValida = False
                            tupla_consulta = (dia_consulta, dia_consulta, año_consulta)
                    print("-"*100)
                    print(f"**\t\tREPORTE DE RESERVACIONES PARA EL DIA {fecha_valida}\t\t**")
                    print("-"*100)
                    print("{:<20} {:<20} {:<20} {:<20}".format('Sala','Cliente', 'Evento', 'Turno'))
                    print("-"*100)
                    
                    
                    try:
                        with sqlite3.connect("bd_eventos.db") as conn:
                            mi_cursor = conn.cursor()
                            valores_consulta={"fecha_consulta":fecha_valida}
                            mi_cursor.execute("SELECT cve_sala, cve_cliente, nombre, turno FROM reservaciones where fecha_ev = :fecha_consulta", valores_consulta)
                            registro = mi_cursor.fetchall()
                            if registro:
                                for clave_c, clave_c, evento, turno in registro:
                                    print("{:<20} {:<20} {:<20} {:<20}".format(clave_c, clave_c, evento, turno))
                            else:
                                print(f"No se encontraron reservaciones para el dia: {fecha_consulta}")
                            
                    except Error as e:
                        print (e)
                    except:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                    print("")
                    print("-"*42+ "FIN DEL REPORTE"+ "-"*43)
                    fechaExistente = True
                
                if opcion2 == 2:
                    print("Reporte en Excel\n")
                    fechaExistente=False
                    fecha_consulta = input("Ingrese la fecha que desea consultar (dd/mm/aaaa): ")
                    fecha_consulta = datetime.datetime.strptime(fecha_consulta,"%d/%m/%Y").date()
                    dia_consulta = fecha_consulta.day
                    mes_consulta = fecha_consulta.month
                    año_consulta = fecha_consulta.year            
                    tupla_consulta = (dia_consulta, mes_consulta, año_consulta)      
                      
                    libro = Workbook()
                    hoja = libro.active
                    hoja["A1"].value = "REPORTE DE EVENTOS PARA EL DIA: "
                    hoja["B1"].value = fecha_consulta
                    hoja["A2"].value= "SALA"
                    hoja["B2"].value = "NOMBRE CLIENTE"
                    hoja["C2"].value = "APELLIDO CLIENTE"
                    hoja["D2"].value = "EVENTO"
                    hoja["E2"].value = "TURNO"
                    
                    try:
                        with sqlite3.connect("bd_eventos.db") as conn:
                            mi_cursor = conn.cursor()
                            valores_consulta={"fecha_consulta":fecha_consulta}
                            mi_cursor.execute("SELECT cve_sala, clientes.nombre, clientes.apellidos, reservaciones.nombre, turno FROM reservaciones INNER JOIN clientes ON reservaciones.cve_cliente=clientes.clave_c where fecha_ev = :fecha_consulta", valores_consulta)
                            registro = mi_cursor.fetchall()
                            if registro:
                                print("Reporte exportado a excel, revise su bandeja de archivos")
                                for clave_sala, nombre_cliente, apellido_cliente, nombre_evento, turno in registro:
                                    row=row+1
                                    
                                    hoja.cell(row=row+1, column=1).value=clave_sala
                                    hoja.cell(row=row+1, column=2).value=nombre_cliente
                                    hoja.cell(row=row+1, column=3).value=apellido_cliente
                                    hoja.cell(row=row+1, column=4).value=nombre_evento
                                    hoja.cell(row=row+1, column=5).value=turno


                            else:
                                print(f"No se encontraron reservaciones para el dia: {fecha_consulta}")
                            
                    except Error as e:
                        print (e)
                    except:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                        
                    fechaExistente = True
                    file_exists = os.path.exists(f'Consulta_eventos_{fecha_consulta}.xlsx')
                    if file_exists:
                        os.remove(f'Consulta_eventos_{fecha_consulta}.xlsx')                    
                    libro.save(f'Consulta_eventos_{fecha_consulta}.xlsx')                            
                    
                if opcion2 == 3:
                    print("-"*40)
                    opcion = 0
            else:
                opcion = menu()                    
                if opcion == 3:
                    print("Registrar un nuevo cliente\n")
                    while True:
                        while True:
                            nombre_cliente=input("Ingrese el nombre del cliente: ")
                            nombre_sinespacio= nombre_cliente.strip()
                            if nombre_sinespacio == "":
                                print("El nombre del cliente no puede omitirse\n")
                            else:
                                break
                        while True:
                            apellidos=input("Ingrese el apellido del cliente: ")
                            apellido_sinespacio= apellidos.strip()
                            if apellido_sinespacio == "":
                                print("El apellido no puede quedar vacio ")
                            else:
                                print("Cliente agregado.\n")
                                try:
                                    with sqlite3.connect("bd_eventos.db") as conn:
                                        mi_cursor = conn.cursor()
                                        valores = {"nombre":nombre_sinespacio, "apellidos":apellido_sinespacio}
                                        mi_cursor.execute("INSERT INTO clientes (nombre, apellidos) VALUES(:nombre,:apellidos)", valores)
                                except Error as e:
                                    print (e)
                                except:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                finally:
                                    conn.close()
                                    break
                        break
                    
                if opcion == 4:
                    print("Registrar una sala\n")
                    while True:
                        while True:
                            nombre_sala = input("Ingrese el nombre de la sala: ")
                            nombre_sinespaciosala= nombre_sala.strip()
                            if nombre_sinespaciosala == "":
                                print("El nombre de la sala no debe omitirse\n")
                            else:
                                break
                        while True:
                            cupo_correcto = True
                            try:
                                cupo_sala = int(input("Ingrese el cupo de la sala: "))
                                if cupo_sala <= 0:
                                    print("El cupo de la sala debe ser un numero mayor a 0\n")
                                    cupo_correcto = False
                                else:
                                    try:
                                        with sqlite3.connect("bd_eventos.db") as conn:
                                            mi_cursor = conn.cursor()
                                            valores = {"nombre":nombre_sinespaciosala, "cupo":cupo_sala}
                                            mi_cursor.execute("INSERT INTO salas (nombre, cupo) VALUES(:nombre,:cupo)", valores)
                                    except Error as e:
                                        print (e)
                                    except:
                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    finally:
                                        conn.close()
                                        print("Sala agregada.\n")
                                        break
                            except ValueError:
                                print("El dato ingresado no es válido, ingrese un cupo correcto")
                                cupo_correcto = False
                            if cupo_correcto:
                                break
                        break
                if opcion == 5:
                    print("Usted a salido con éxito\n")
                    break






































